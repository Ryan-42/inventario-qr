"""
Testes: endpoint GET /metricas, aba de métricas no Excel final e SELECT FOR UPDATE.
"""
from __future__ import annotations
import io
import threading


def _reg(client, sid, codigo, qtd, operador="Op1"):
    return client.post(
        f"/api/sessoes/{sid}/contagens",
        json={"codigo": codigo, "quantidade_encontrada": qtd, "operador": operador},
    )


# ── P2.8: Endpoint /metricas ────────────────────────────────────────────────

def test_metricas_sessao_vazia(client, sessao):
    """Sessão sem itens retorna estrutura coerente com zeros."""
    r = client.get(f"/api/sessoes/{sessao['id']}/metricas")
    assert r.status_code == 200
    data = r.json()
    assert data["total_itens"] == 0
    assert data["total_contagens_atuais"] == 0
    assert data["total_tentativas_historico"] == 0
    assert data["taxa_divergencia_pct"] == 0.0
    assert data["taxa_retrabalho_pct"] == 0.0
    assert data["pct_rastreabilidade"] == 0.0


def test_metricas_campos_obrigatorios(client, sessao_com_itens):
    """Resposta deve conter todos os campos de KPI."""
    sid = sessao_com_itens["id"]
    _reg(client, sid, "ABC-001", 10)
    r = client.get(f"/api/sessoes/{sid}/metricas")
    assert r.status_code == 200
    data = r.json()
    campos = [
        "sessao_id", "sessao_codigo", "sessao_nome", "status",
        "inicio", "duracao_minutos",
        "total_itens", "total_contagens_atuais", "total_tentativas_historico",
        "itens_por_minuto",
        "taxa_divergencia_pct", "divergencias_absolutas",
        "taxa_retrabalho_pct", "retrabalho_absoluto",
        "pct_rastreabilidade", "contagens_com_operador",
        "por_operador",
    ]
    for campo in campos:
        assert campo in data, f"Campo '{campo}' ausente na resposta de métricas"


def test_metricas_sem_divergencia(client, sessao_com_itens):
    """Todos os itens OK → taxa_divergencia=0, taxa_retrabalho=0."""
    sid = sessao_com_itens["id"]
    _reg(client, sid, "ABC-001", 10)   # base=10 OK
    _reg(client, sid, "ABC-002", 5)    # base=5  OK
    _reg(client, sid, "ABC-003", 20)   # base=20 OK

    r = client.get(f"/api/sessoes/{sid}/metricas")
    assert r.status_code == 200
    data = r.json()
    assert data["taxa_divergencia_pct"] == 0.0
    assert data["divergencias_absolutas"] == 0
    assert data["taxa_retrabalho_pct"] == 0.0
    assert data["retrabalho_absoluto"] == 0
    assert data["total_contagens_atuais"] == 3
    assert data["total_tentativas_historico"] == 3


def test_metricas_com_divergencia(client, sessao_com_itens):
    """1 item divergente em 3 → taxa_divergencia ≈ 33%."""
    sid = sessao_com_itens["id"]
    _reg(client, sid, "ABC-001", 10)
    _reg(client, sid, "ABC-002", 3)    # divergente (base=5)
    _reg(client, sid, "ABC-003", 20)

    r = client.get(f"/api/sessoes/{sid}/metricas")
    data = r.json()
    assert data["divergencias_absolutas"] == 1
    assert abs(data["taxa_divergencia_pct"] - 33.33) < 0.1


def test_metricas_retrabalho(client, sessao_com_itens):
    """ABC-002 recontado 2x → retrabalho_absoluto=1."""
    sid = sessao_com_itens["id"]
    _reg(client, sid, "ABC-001", 10)
    _reg(client, sid, "ABC-002", 3)    # divergente R1
    _reg(client, sid, "ABC-003", 20)
    _reg(client, sid, "ABC-002", 3)    # recontagem R2 — confirma PARA_AJUSTE

    r = client.get(f"/api/sessoes/{sid}/metricas")
    data = r.json()
    # histórico: 4 entradas para 3 itens únicos → retrabalho_absoluto = 4-3 = 1
    assert data["total_tentativas_historico"] == 4
    assert data["retrabalho_absoluto"] == 1
    assert data["taxa_retrabalho_pct"] > 0


def test_metricas_rastreabilidade(client, sessao_com_itens):
    """Apenas contagens com operador contam para rastreabilidade."""
    sid = sessao_com_itens["id"]
    # 2 com operador, 1 sem
    _reg(client, sid, "ABC-001", 10, operador="Op1")
    _reg(client, sid, "ABC-002", 5,  operador="Op2")
    client.post(
        f"/api/sessoes/{sid}/contagens",
        json={"codigo": "ABC-003", "quantidade_encontrada": 20},  # sem operador
    )

    r = client.get(f"/api/sessoes/{sid}/metricas")
    data = r.json()
    assert data["contagens_com_operador"] == 2
    assert abs(data["pct_rastreabilidade"] - 66.67) < 0.1


def test_metricas_por_operador(client, sessao_com_itens):
    """Breakdown por operador deve listar cada operador com suas métricas."""
    sid = sessao_com_itens["id"]
    _reg(client, sid, "ABC-001", 10, operador="Ana")
    _reg(client, sid, "ABC-002", 5,  operador="Bruno")
    _reg(client, sid, "ABC-003", 20, operador="Ana")

    r = client.get(f"/api/sessoes/{sid}/metricas")
    data = r.json()
    ops = {op["operador"]: op for op in data["por_operador"]}
    assert "Ana" in ops
    assert "Bruno" in ops
    assert ops["Ana"]["contagens"] == 2
    assert ops["Ana"]["itens_unicos"] == 2
    assert ops["Bruno"]["contagens"] == 1


def test_metricas_sessao_inexistente(client):
    """Sessão não encontrada retorna 404."""
    r = client.get("/api/sessoes/nao-existe/metricas")
    assert r.status_code == 404


# ── P2.9: Aba de métricas no Excel final ─────────────────────────────────────

def test_excel_final_tem_aba_metricas(client, sessao_com_itens):
    """relatorio-final-excel deve incluir aba 'Métricas Produtividade'."""
    import openpyxl

    sid = sessao_com_itens["id"]
    tok = sessao_com_itens["token_admin"]
    _reg(client, sid, "ABC-001", 10)
    _reg(client, sid, "ABC-002", 3)    # divergente

    r = client.post(
        f"/api/sessoes/{sid}/exportar/relatorio-final-excel",
        json={"token_admin": tok},
    )
    assert r.status_code == 200, r.text
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "Métricas Produtividade" in wb.sheetnames, f"Abas: {wb.sheetnames}"


def test_excel_final_metricas_conteudo(client, sessao_com_itens):
    """Aba de métricas deve conter as linhas esperadas de KPI."""
    import openpyxl

    sid = sessao_com_itens["id"]
    tok = sessao_com_itens["token_admin"]
    _reg(client, sid, "ABC-001", 10, operador="Ana")
    _reg(client, sid, "ABC-002", 5,  operador="Bruno")
    _reg(client, sid, "ABC-003", 20, operador="Ana")

    r = client.post(
        f"/api/sessoes/{sid}/exportar/relatorio-final-excel",
        json={"token_admin": tok},
    )
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["Métricas Produtividade"]
    metricas_labels = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
    assert any("Divergência" in (lbl or "") for lbl in metricas_labels)
    assert any("Rastreabilidade" in (lbl or "") for lbl in metricas_labels)
    assert any("Retrabalho" in (lbl or "") for lbl in metricas_labels)


def test_excel_final_tem_aba_produtividade_por_operador(client, sessao_com_itens):
    """Aba 'Produtividade por Operador' deve aparecer quando há histórico."""
    import openpyxl

    sid = sessao_com_itens["id"]
    tok = sessao_com_itens["token_admin"]
    _reg(client, sid, "ABC-001", 10, operador="Ana")
    _reg(client, sid, "ABC-002", 5,  operador="Bruno")

    r = client.post(
        f"/api/sessoes/{sid}/exportar/relatorio-final-excel",
        json={"token_admin": tok},
    )
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "Produtividade por Operador" in wb.sheetnames, f"Abas: {wb.sheetnames}"
    ws = wb["Produtividade por Operador"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "Operador" in headers
    assert "Itens/min" in headers


# ── P0.3: SELECT FOR UPDATE — comportamento com contagem existente ───────────

def test_contagem_atualizada_preserva_historico_completo(client, sessao_com_itens):
    """Recontagem de mesmo item: histórico deve registrar todas as tentativas."""
    sid = sessao_com_itens["id"]
    _reg(client, sid, "ABC-002", 3, operador="Op1")  # R1 divergente
    _reg(client, sid, "ABC-002", 4, operador="Op2")  # R2 qtd diferente

    r = client.get(f"/api/sessoes/{sid}/historico?codigo=ABC-002")
    assert r.status_code == 200
    hist = r.json()
    assert len(hist) == 2, f"Esperado 2 entradas, got {len(hist)}: {hist}"
    qtds = {h["quantidade_encontrada"] for h in hist}
    assert qtds == {3, 4}
    ops = {h["operador"] for h in hist}
    assert ops == {"Op1", "Op2"}


def test_contagem_concorrente_sequencial_nao_perde_historico(client, sessao_com_itens):
    """
    Simula dois updates sequenciais ao mesmo item: ambos devem ser registrados no histórico.
    SQLite serializa writes, portanto este teste é determinístico mesmo sem FOR UPDATE.
    O comportamento com PostgreSQL + FOR UPDATE é idêntico — valida a lógica de negócio.
    """
    sid = sessao_com_itens["id"]
    errors = []

    def contar(qtd, op):
        try:
            r = _reg(client, sid, "ABC-001", qtd, operador=op)
            if r.status_code not in (200, 201, 409):
                errors.append(f"status inesperado {r.status_code}: {r.text}")
        except Exception as exc:
            errors.append(str(exc))

    # Envia sequencialmente (SQLite não suporta concorrência real em testes)
    contar(10, "OperadorA")
    contar(11, "OperadorB")  # qtd diferente → avança rodada

    assert not errors, f"Erros: {errors}"
    # Estado final deve estar consistente
    r = client.get(f"/api/sessoes/{sid}/historico?codigo=ABC-001")
    hist = r.json()
    assert len(hist) == 2
    assert all(h["codigo"] == "ABC-001" for h in hist)


def test_codigo_inexistente_retorna_404_nao_crash(client, sessao_com_itens):
    """Leitura de código inexistente deve retornar 404, não 500."""
    sid = sessao_com_itens["id"]
    r = client.post(
        f"/api/sessoes/{sid}/contagens",
        json={"codigo": "NAO-EXISTE-9999", "quantidade_encontrada": 1},
    )
    assert r.status_code == 404
    assert "não encontrado" in r.json()["detail"].lower()


def test_codigo_malformado_retorna_422(client, sessao_com_itens):
    """Código vazio deve falhar na validação do schema."""
    sid = sessao_com_itens["id"]
    r = client.post(
        f"/api/sessoes/{sid}/contagens",
        json={"codigo": "   ", "quantidade_encontrada": 1},
    )
    assert r.status_code == 422
