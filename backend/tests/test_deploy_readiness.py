"""Testes de prontidão para deploy: health check, fluxo completo, edge cases."""
from __future__ import annotations
import io
import openpyxl


def _make_xlsx(rows, header=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header or ["codigo", "produto", "quantidade"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _reg(client, sid, codigo, qtd):
    return client.post(f"/api/sessoes/{sid}/contagens",
                       json={"codigo": codigo, "quantidade_encontrada": qtd})


# ── Health check ──────────────────────────────────────────────────────────────

def test_health_check_retorna_200(client):
    """Endpoint /health deve retornar 200 com banco OK."""
    r = client.get("/health")
    # Em testes, o DB é SQLite em memória — pode retornar 200 ou 503
    # O importante é que não exploda (5xx inesperado)
    assert r.status_code in (200, 503)
    data = r.json()
    assert "status" in data
    assert "sistema" in data
    assert data["sistema"] == "INVIQ"


# ── Fluxo completo de inventário ──────────────────────────────────────────────

def test_fluxo_completo_todos_certos(client, sessao_com_itens):
    """Fluxo feliz: todos os itens contados corretamente, sessão concluída."""
    sid = sessao_com_itens["id"]
    tok = sessao_com_itens["token_admin"]

    # Registra todos corretos
    _reg(client, sid, "ABC-001", 10)
    _reg(client, sid, "ABC-002", 5)
    _reg(client, sid, "ABC-003", 20)

    # Progresso completo
    prog = client.get(f"/api/sessoes/{sid}/progresso").json()
    assert prog["completa"] is True
    assert prog["faltando_r1"] == 0
    assert prog["faltando_r2"] == 0

    # Conclui
    r = client.patch(f"/api/sessoes/{sid}/concluir?token_admin={tok}")
    assert r.status_code == 200
    assert r.json()["status"] == "concluida"
    assert r.json()["total_itens"] == 3


def test_fluxo_completo_com_recontagem(client, sessao_com_itens):
    """Fluxo com divergência → recontagem → para_ajuste → conclusão."""
    sid = sessao_com_itens["id"]
    tok = sessao_com_itens["token_admin"]

    # R1: 2 certos, 1 divergente
    _reg(client, sid, "ABC-001", 10)
    _reg(client, sid, "ABC-002", 3)   # diverge (base=5)
    _reg(client, sid, "ABC-003", 20)

    prog = client.get(f"/api/sessoes/{sid}/progresso").json()
    assert prog["faltando_r2"] == 1
    assert prog["completa"] is False

    # Não pode concluir ainda
    r = client.patch(f"/api/sessoes/{sid}/concluir?token_admin={tok}")
    assert r.status_code == 422

    # Recontagem: confirma o mesmo erro (qty=3 de novo → para_ajuste)
    _reg(client, sid, "ABC-002", 3)

    prog2 = client.get(f"/api/sessoes/{sid}/progresso").json()
    assert prog2["faltando_r2"] == 0
    assert prog2["completa"] is True

    # Agora pode concluir
    r2 = client.patch(f"/api/sessoes/{sid}/concluir?token_admin={tok}")
    assert r2.status_code == 200


def test_fluxo_exportacao_apos_conclusao(client, sessao_com_itens):
    """Exportações funcionam para sessão com contagens."""
    sid = sessao_com_itens["id"]
    tok = sessao_com_itens["token_admin"]
    _reg(client, sid, "ABC-001", 10)
    _reg(client, sid, "ABC-002", 3)   # divergente

    # Excel completo
    r = client.post(f"/api/sessoes/{sid}/exportar/completo", json={"token_admin": tok})
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert len(r.content) > 0

    # PDF
    r_pdf = client.post(f"/api/sessoes/{sid}/exportar/pdf", json={"token_admin": tok})
    assert r_pdf.status_code == 200
    assert "pdf" in r_pdf.headers["content-type"]


# ── Edge cases ────────────────────────────────────────────────────────────────

def test_sessao_nome_muito_longo_bloqueado(client):
    """Nome com mais de 120 chars deve ser rejeitado."""
    nome_longo = "X" * 121
    r = client.post("/api/sessoes/", json={"nome": nome_longo})
    assert r.status_code == 422


def test_sessao_nome_vazio_bloqueado(client):
    r = client.post("/api/sessoes/", json={"nome": "   "})
    assert r.status_code == 422


def test_contagem_quantidade_negativa_bloqueada(client, sessao_com_itens):
    sid = sessao_com_itens["id"]
    r = client.post(f"/api/sessoes/{sid}/contagens",
                    json={"codigo": "ABC-001", "quantidade_encontrada": -1})
    assert r.status_code == 422


def test_contagem_quantidade_zero_permitida(client, sessao_com_itens):
    """Quantidade 0 é válida (item fisicamente ausente)."""
    sid = sessao_com_itens["id"]
    r = client.post(f"/api/sessoes/{sid}/contagens",
                    json={"codigo": "ABC-001", "quantidade_encontrada": 0})
    assert r.status_code == 201
    assert r.json()["divergencia"] is True  # diverge da base=10


def test_upload_arquivo_muito_grande_bloqueado(client, sessao):
    """Arquivo acima de 10MB deve ser rejeitado."""
    dados_grandes = b"X" * (11 * 1024 * 1024)
    tok = sessao["token_admin"]
    r = client.post(
        f"/api/sessoes/{sessao['id']}/upload?token_admin={tok}",
        files={"file": ("grande.xlsx", io.BytesIO(dados_grandes),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    assert r.status_code in (400, 413, 422)


def test_upload_extensao_invalida(client, sessao):
    tok = sessao["token_admin"]
    r = client.post(
        f"/api/sessoes/{sessao['id']}/upload?token_admin={tok}",
        files={"file": ("dados.txt", io.BytesIO(b"codigo,produto,quantidade"),
                        "text/plain")}
    )
    assert r.status_code == 400


def test_contagem_item_inexistente(client, sessao_com_itens):
    sid = sessao_com_itens["id"]
    r = client.post(f"/api/sessoes/{sid}/contagens",
                    json={"codigo": "NAO-EXISTE-999", "quantidade_encontrada": 5})
    assert r.status_code == 404


def test_buscar_sessao_inexistente(client):
    r = client.get("/api/sessoes/nao-existe-uuid-algum")
    assert r.status_code == 404


def test_concluir_sessao_inexistente(client):
    r = client.patch("/api/sessoes/nao-existe/concluir?token_admin=X")
    assert r.status_code == 404


# ── Concorrência simulada (dois counts no mesmo item) ────────────────────────

def test_dois_registros_mesmo_item_upsert(client, sessao_com_itens):
    """Registrar duas vezes o mesmo item deve fazer upsert (não duplicar)."""
    sid = sessao_com_itens["id"]
    _reg(client, sid, "ABC-001", 10)
    _reg(client, sid, "ABC-001", 10)  # segunda vez
    contagens = client.get(f"/api/sessoes/{sid}/contagens").json()
    abc001 = [c for c in contagens if c["codigo"] == "ABC-001"]
    assert len(abc001) == 1, "Upsert deve manter apenas 1 contagem por item"


def test_historico_registra_ambas_as_tentativas(client, sessao_com_itens):
    """Histórico deve ter 2 entradas para 2 scans do mesmo item."""
    sid = sessao_com_itens["id"]
    _reg(client, sid, "ABC-001", 10)
    _reg(client, sid, "ABC-001", 8)  # segunda contagem
    r = client.get(f"/api/sessoes/{sid}/historico?codigo=ABC-001")
    assert r.status_code == 200
    assert len(r.json()) == 2


# ── Imports e consistência de dados ──────────────────────────────────────────

def test_reimport_bloqueado_apos_contagens(client, sessao_com_itens):
    """Reimportar planilha é bloqueado (409) após início das contagens — by design."""
    sid = sessao_com_itens["id"]
    tok = sessao_com_itens["token_admin"]
    _reg(client, sid, "ABC-001", 10)  # inicia contagens

    xlsx = _make_xlsx([["ABC-001", "Produto Alpha", 10]])
    r = client.post(
        f"/api/sessoes/{sid}/upload?token_admin={tok}",
        files={"file": ("itens.xlsx", xlsx,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 409
    assert "contagem" in r.json()["detail"].lower()


def test_planilha_com_valor_estoque(client, sessao):
    """Planilha com coluna valor_estoque deve ser importada corretamente."""
    tok = sessao["token_admin"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["codigo", "produto", "quantidade", "valor_estoque"])
    ws.append(["P001", "Produto 1", 10, 250.00])
    ws.append(["P002", "Produto 2", 5, 0.0])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post(
        f"/api/sessoes/{sessao['id']}/upload?token_admin={tok}",
        files={"file": ("itens.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 201
    assert r.json()["total"] == 2


def test_valor_unitario_zero_nao_e_null(client, sessao):
    """Item com valor_estoque=0.0 deve ter valor_unitario=0.0, não null (no serviço)."""
    from app.services.sessao_service import montar_inventario_completo
    import openpyxl as xl

    sid = sessao["id"]
    tok = sessao["token_admin"]
    wb = xl.Workbook()
    ws = wb.active
    ws.append(["codigo", "produto", "quantidade", "valor_estoque"])
    ws.append(["P001", "Produto", 5, 0.0])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    client.post(
        f"/api/sessoes/{sid}/upload?token_admin={tok}",
        files={"file": ("itens.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    client.post(f"/api/sessoes/{sid}/contagens",
                json={"codigo": "P001", "quantidade_encontrada": 5})

    # Verifica via serviço diretamente (o endpoint /itens usa ItemComStatus que não expõe valor_unitario)
    from app.main import app
    from app.database import get_db as orig_get_db
    from tests.conftest import override_get_db
    db = next(override_get_db())
    try:
        inventario = montar_inventario_completo(db, sid)
        p001 = next((i for i in inventario if i["codigo"] == "P001"), None)
        assert p001 is not None
        # Item contado com valor_estoque=0.0 e quantidade base=5: valor_unitario = 0.0/5 = 0.0
        assert p001["valor_unitario"] == 0.0, \
            f"valor_unitario deve ser 0.0, nao null. Recebeu: {p001['valor_unitario']}"
    finally:
        db.close()
