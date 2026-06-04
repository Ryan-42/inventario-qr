"""Testes: upload de planilha e busca de itens."""
from __future__ import annotations

import io
import openpyxl


def _make_xlsx(rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["codigo", "produto", "quantidade"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, sessao, xlsx, filename="itens.xlsx"):
    tok = sessao["token_admin"]
    return client.post(
        f"/api/sessoes/{sessao['id']}/upload?token_admin={tok}",
        files={"file": (filename, xlsx,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def test_upload_planilha_sucesso(client, sessao):
    xlsx = _make_xlsx([["P001", "Produto Um", 10], ["P002", "Produto Dois", 5]])
    r = _upload(client, sessao, xlsx)
    assert r.status_code == 201
    assert r.json()["total"] == 2


def test_upload_planilha_sessao_inexistente(client):
    xlsx = _make_xlsx([["P001", "X", 1]])
    r = client.post(
        "/api/sessoes/nao-existe/upload?token_admin=X",
        files={"file": ("itens.xlsx", xlsx, "application/vnd.ms-excel")},
    )
    assert r.status_code == 404


def test_upload_extensao_invalida(client, sessao):
    tok = sessao["token_admin"]
    r = client.post(
        f"/api/sessoes/{sessao['id']}/upload?token_admin={tok}",
        files={"file": ("itens.txt", b"dados", "text/plain")},
    )
    assert r.status_code == 400


def test_upload_arquivo_vazio(client, sessao):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["codigo", "produto", "quantidade"])
    buf = io.BytesIO()
    wb.save(buf)
    r = _upload(client, sessao, buf.getvalue(), "vazio.xlsx")
    assert r.status_code == 422


def test_upload_colunas_faltando(client, sessao):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["codigo", "quantidade"])  # falta 'produto'
    ws.append(["P001", 10])
    buf = io.BytesIO()
    wb.save(buf)
    r = _upload(client, sessao, buf.getvalue(), "sem_produto.xlsx")
    assert r.status_code in (201, 422)


def test_upload_token_invalido_retorna_403(client, sessao):
    xlsx = _make_xlsx([["P001", "Produto", 10]])
    r = client.post(
        f"/api/sessoes/{sessao['id']}/upload?token_admin=ERRADO",
        files={"file": ("itens.xlsx", xlsx, "application/vnd.ms-excel")},
    )
    assert r.status_code == 403


def test_buscar_item_existente(client, sessao_com_itens):
    sid = sessao_com_itens["id"]
    r = client.get(f"/api/sessoes/{sid}/buscar/ABC-001")
    assert r.status_code == 200
    data = r.json()
    assert data["codigo"] == "ABC-001"
    assert data["produto"] == "Produto Alpha"
    assert data["quantidade_base"] == 10
    assert data["ja_contado"] is False
    assert data["contagem_anterior"] is None


def test_buscar_item_inexistente(client, sessao_com_itens):
    sid = sessao_com_itens["id"]
    r = client.get(f"/api/sessoes/{sid}/buscar/NAO-EXISTE")
    assert r.status_code == 404


def test_listar_itens(client, sessao_com_itens):
    sid = sessao_com_itens["id"]
    r = client.get(f"/api/sessoes/{sid}/itens")
    assert r.status_code == 200
    itens = r.json()
    assert len(itens) == 3
    codigos = {i["codigo"] for i in itens}
    assert codigos == {"ABC-001", "ABC-002", "ABC-003"}


def test_validar_planilha_ok(client, sessao):
    xlsx = _make_xlsx([["P001", "Produto Um", 10]])
    r = client.post(
        f"/api/sessoes/{sessao['id']}/validar-planilha",
        files={"file": ("itens.xlsx", xlsx, "application/vnd.ms-excel")},
    )
    assert r.status_code == 200
    data = r.json()
    assert "valido" in data
    assert "total_validos" in data


def test_validar_planilha_csv(client, sessao):
    csv_data = b"codigo,produto,quantidade\nP001,Produto,10\n"
    r = client.post(
        f"/api/sessoes/{sessao['id']}/validar-planilha",
        files={"file": ("itens.csv", csv_data, "text/csv")},
    )
    assert r.status_code in (200, 400)


def test_upload_bloqueado_sessao_concluida(client, sessao_com_itens):
    sid = sessao_com_itens["id"]
    for codigo, qtd in [("ABC-001", 10), ("ABC-002", 5), ("ABC-003", 20)]:
        client.post(f"/api/sessoes/{sid}/contagens",
                    json={"codigo": codigo, "quantidade_encontrada": qtd})
    tok = sessao_com_itens["token_admin"]
    r_concluir = client.patch(f"/api/sessoes/{sid}/concluir?token_admin={tok}")
    assert r_concluir.status_code == 200, f"Falha ao concluir: {r_concluir.text}"

    xlsx = _make_xlsx([["P001", "Produto", 10]])
    r = client.post(
        f"/api/sessoes/{sid}/upload?token_admin={tok}",
        files={"file": ("itens.xlsx", xlsx, "application/vnd.ms-excel")},
    )
    assert r.status_code == 409
    assert "concluida" in r.json()["detail"]


def test_upload_bloqueado_apos_contagens(client, sessao_com_itens):
    sid = sessao_com_itens["id"]
    tok = sessao_com_itens["token_admin"]
    client.post(f"/api/sessoes/{sid}/contagens",
                json={"codigo": "ABC-001", "quantidade_encontrada": 10})
    xlsx = _make_xlsx([["P001", "Novo Produto", 10]])
    r = client.post(
        f"/api/sessoes/{sid}/upload?token_admin={tok}",
        files={"file": ("itens.xlsx", xlsx, "application/vnd.ms-excel")},
    )
    assert r.status_code == 409
    assert "contagem" in r.json()["detail"].lower()
