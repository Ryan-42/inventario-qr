"""
Testes: edge cases de geração de PDF — foco em situações que causavam crashes.
"""
from __future__ import annotations

import io
import openpyxl


def _registrar(client, sessao_id, codigo, qtd, operador="Tester"):
    return client.post(
        f"/api/sessoes/{sessao_id}/contagens",
        json={"codigo": codigo, "quantidade_encontrada": qtd, "operador": operador},
    )


def _export(client, sid, tok, path):
    return client.post(f"/api/sessoes/{sid}/exportar/{path}", json={"token_admin": tok})


def _sessao_com_itens_especiais(client, sessao, nomes_itens):
    """Cria sessão com itens cujos nomes podem quebrar o parser XML do ReportLab."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["codigo", "produto", "quantidade"])
    for cod, nome, qtd in nomes_itens:
        ws.append([cod, nome, qtd])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    tok = sessao["token_admin"]
    r = client.post(
        f"/api/sessoes/{sessao['id']}/upload?token_admin={tok}",
        files={"file": ("itens.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 201, f"Upload falhou: {r.text}"
    return sessao


# ── Testes de caracteres especiais que quebravam o ReportLab ─────────────────

def test_pdf_produto_com_angulo_menor_maior(client, sessao):
    """Nomes com < e > não devem crashar o ReportLab."""
    itens = [
        ("P001", "Parafuso M4 <=> M6", 10),
        ("P002", "Cabo USB <Tipo-A>", 5),
    ]
    sessao = _sessao_com_itens_especiais(client, sessao, itens)
    _registrar(client, sessao["id"], "P001", 10)
    _registrar(client, sessao["id"], "P002", 3)

    r = _export(client, sessao["id"], sessao["token_admin"], "pdf")
    assert r.status_code == 200, f"PDF falhou: {r.text}"
    assert r.headers["content-type"] == "application/pdf"
    assert r.content[:4] == b"%PDF"


def test_pdf_produto_com_e_comercial(client, sessao):
    """Nomes com & não devem crashar o ReportLab."""
    itens = [("X001", "Sal & Açúcar Refinado", 20)]
    sessao = _sessao_com_itens_especiais(client, sessao, itens)
    _registrar(client, sessao["id"], "X001", 20)

    r = _export(client, sessao["id"], sessao["token_admin"], "pdf")
    assert r.status_code == 200
    assert r.content[:4] == b"%PDF"


def test_relatorio_final_pdf_produto_com_html_chars(client, sessao):
    """Relatório Final PDF não deve crashar com nomes de produto contendo <, > e &."""
    itens = [
        ("A001", "<Produto> Especial & Único", 10),
        ("A002", "Item com \"aspas\" e 'apóstrofo'", 5),
    ]
    sessao = _sessao_com_itens_especiais(client, sessao, itens)
    _registrar(client, sessao["id"], "A001", 10)
    _registrar(client, sessao["id"], "A002", 3)  # divergente

    r = _export(client, sessao["id"], sessao["token_admin"], "relatorio-final-pdf")
    assert r.status_code == 200, f"Relatório Final PDF falhou: {r.text}"
    assert r.content[:4] == b"%PDF"
    assert len(r.content) > 2000


def test_pdf_nome_sessao_com_caracteres_especiais(client):
    """PDF não deve crashar quando o nome da sessão tem <, > ou &."""
    r = client.post("/api/sessoes/", json={"nome": "Inventário <Depósito A> & B"})
    assert r.status_code == 201
    dados = r.json()
    sid, tok = dados["id"], dados["token_admin"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["codigo", "produto", "quantidade"])
    ws.append(["Z001", "Produto Z", 5])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    client.post(f"/api/sessoes/{sid}/upload?token_admin={tok}",
                files={"file": ("i.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    client.post(f"/api/sessoes/{sid}/contagens", json={"codigo": "Z001", "quantidade_encontrada": 5})

    r = client.post(f"/api/sessoes/{sid}/exportar/pdf", json={"token_admin": tok})
    assert r.status_code == 200
    assert r.content[:4] == b"%PDF"


def test_pdf_sem_contagens_retorna_pdf(client, sessao_com_itens):
    """PDF sem nenhuma contagem registrada deve gerar relatório com itens pendentes."""
    sid = sessao_com_itens["id"]
    tok = sessao_com_itens["token_admin"]

    r = _export(client, sid, tok, "pdf")
    assert r.status_code == 200
    assert r.content[:4] == b"%PDF"


def test_relatorio_final_pdf_tamanho_minimo(client, sessao_com_itens):
    """PDF executivo deve ter tamanho mínimo razoável — verifica que não está vazio/corrompido."""
    sid = sessao_com_itens["id"]
    tok = sessao_com_itens["token_admin"]
    _registrar(client, sid, "ABC-001", 10)
    _registrar(client, sid, "ABC-002", 3)
    _registrar(client, sid, "ABC-003", 20)

    r = _export(client, sid, tok, "relatorio-final-pdf")
    assert r.status_code == 200
    assert len(r.content) > 3_000, f"PDF muito pequeno ({len(r.content)} bytes) — possível arquivo corrompido"


def test_relatorio_final_excel_tamanho_minimo(client, sessao_com_itens):
    """Excel executivo deve ter tamanho mínimo razoável."""
    sid = sessao_com_itens["id"]
    tok = sessao_com_itens["token_admin"]
    _registrar(client, sid, "ABC-001", 10)
    _registrar(client, sid, "ABC-002", 3)

    r = _export(client, sid, tok, "relatorio-final-excel")
    assert r.status_code == 200
    assert len(r.content) > 3_000, f"Excel muito pequeno ({len(r.content)} bytes)"

    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert len(wb.sheetnames) >= 4, f"Excel deveria ter pelo menos 4 abas, tem: {wb.sheetnames}"


def test_pdf_muitos_itens_nao_crashar(client, sessao):
    """PDF com 50 itens não deve crashar — testa paginação do ReportLab."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["codigo", "produto", "quantidade"])
    for i in range(50):
        ws.append([f"ITEM-{i:03d}", f"Produto {i} — Descrição Longa de Teste", 10])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    tok = sessao["token_admin"]
    client.post(f"/api/sessoes/{sessao['id']}/upload?token_admin={tok}",
                files={"file": ("big.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})

    for i in range(50):
        client.post(f"/api/sessoes/{sessao['id']}/contagens",
                    json={"codigo": f"ITEM-{i:03d}", "quantidade_encontrada": 10 if i % 3 != 0 else 7})

    r = client.post(f"/api/sessoes/{sessao['id']}/exportar/pdf", json={"token_admin": tok})
    assert r.status_code == 200
    assert r.content[:4] == b"%PDF"
    assert len(r.content) > 5_000


def test_etiquetas_pdf_com_codigo_especial(client, sessao):
    """Etiquetas com códigos contendo caracteres especiais não devem crashar."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["codigo", "produto", "quantidade"])
    ws.append(["EAN-7891234567890", "Produto com EAN longo", 5])
    ws.append(["COD/ABC/123", "Produto barra", 3])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    tok = sessao["token_admin"]
    client.post(f"/api/sessoes/{sessao['id']}/upload?token_admin={tok}",
                files={"file": ("e.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})

    r = client.post(f"/api/sessoes/{sessao['id']}/exportar/etiquetas", json={"token_admin": tok})
    assert r.status_code == 200
    assert r.content[:4] == b"%PDF"
