"""Testes unitários: excel_service (sem FastAPI, sem banco)."""
from __future__ import annotations

import io
import pytest
import openpyxl
from fastapi import HTTPException

from app.services.excel_service import importar_planilha, exportar_inventario_completo, exportar_divergencias


def _make_xlsx(rows: list[list], header=None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header or ["codigo", "produto", "quantidade"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── importar_planilha ────────────────────────────────────────────────────────

def test_importar_planilha_basica():
    xlsx = _make_xlsx([["A001", "Produto Alpha", 10], ["A002", "Produto Beta", 5]])
    itens = importar_planilha(xlsx)
    assert len(itens) == 2
    assert itens[0] == {"codigo": "A001", "produto": "Produto Alpha", "quantidade_base": 10}
    assert itens[1] == {"codigo": "A002", "produto": "Produto Beta", "quantidade_base": 5}


def test_importar_planilha_colunas_maiusculas():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["CODIGO", "PRODUTO", "QUANTIDADE"])
    ws.append(["X001", "Produto X", 3])
    buf = io.BytesIO()
    wb.save(buf)
    itens = importar_planilha(buf.getvalue())
    assert len(itens) == 1


def test_importar_planilha_linhas_nulas_ignoradas():
    xlsx = _make_xlsx([["A001", "Produto A", 10], [None, None, None], ["A002", "Produto B", 5]])
    itens = importar_planilha(xlsx)
    assert len(itens) == 2


def test_importar_planilha_vazia_levanta_erro():
    xlsx = _make_xlsx([])
    with pytest.raises(HTTPException) as exc:
        importar_planilha(xlsx)
    assert exc.value.status_code == 422


def test_importar_planilha_coluna_faltando():
    xlsx = _make_xlsx([["A001", 10]], header=["codigo", "quantidade"])
    with pytest.raises(HTTPException) as exc:
        importar_planilha(xlsx)
    assert exc.value.status_code == 422


def test_importar_planilha_quantidade_invalida():
    xlsx = _make_xlsx([["A001", "Produto", "nao-e-numero"]])
    with pytest.raises(HTTPException) as exc:
        importar_planilha(xlsx)
    assert exc.value.status_code == 422


def test_importar_arquivo_corrompido():
    with pytest.raises(HTTPException) as exc:
        importar_planilha(b"dados invalidos nao sao xlsx")
    assert exc.value.status_code == 400


# ── exportar_inventario_completo ─────────────────────────────────────────────

def test_exportar_inventario_completo():
    itens = [
        {"codigo": "A001", "produto": "Produto A", "quantidade_base": 10,
         "quantidade_encontrada": 10, "diferenca": 0, "status": "OK"},
        {"codigo": "A002", "produto": "Produto B", "quantidade_base": 5,
         "quantidade_encontrada": 3, "diferenca": -2, "status": "Divergente"},
    ]
    resultado = exportar_inventario_completo(itens)
    assert isinstance(resultado, bytes)
    assert len(resultado) > 0
    # Verifica que é um xlsx válido
    wb = openpyxl.load_workbook(io.BytesIO(resultado))
    assert "Inventário Completo" in wb.sheetnames


def test_exportar_divergencias_vazio():
    resultado = exportar_divergencias([])
    assert isinstance(resultado, bytes)
    wb = openpyxl.load_workbook(io.BytesIO(resultado))
    ws = wb["Divergências"]
    assert ws.max_row == 1  # só cabeçalho


def test_exportar_divergencias_com_dados():
    itens = [
        {"codigo": "A002", "produto": "Produto B", "quantidade_base": 5,
         "quantidade_encontrada": 3, "diferenca": -2},
    ]
    resultado = exportar_divergencias(itens)
    wb = openpyxl.load_workbook(io.BytesIO(resultado))
    ws = wb["Divergências"]
    assert ws.max_row == 2  # cabeçalho + 1 linha
