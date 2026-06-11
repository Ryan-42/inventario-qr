"""
Gera o relatório final de sessão: PDF executivo + Excel detalhado.
Inclui erros, acertos, orientações de melhoria e impacto financeiro.
"""
from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any

import pandas as pd
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
    KeepTogether,
)

# ── Palette ──────────────────────────────────────────────────────────────────
_BLUE       = colors.HexColor("#2563EB")
_BLUE_DARK  = colors.HexColor("#1E40AF")
_BLUE_LIGHT = colors.HexColor("#EFF6FF")
_GREEN      = colors.HexColor("#16A34A")
_GREEN_LIGHT= colors.HexColor("#F0FDF4")
_RED        = colors.HexColor("#DC2626")
_RED_LIGHT  = colors.HexColor("#FEF2F2")
_AMBER      = colors.HexColor("#D97706")
_AMBER_LIGHT= colors.HexColor("#FFFBEB")
_PURPLE     = colors.HexColor("#7C3AED")
_PURPLE_LIGHT=colors.HexColor("#F5F3FF")
_SLATE_900  = colors.HexColor("#0F172A")
_SLATE_700  = colors.HexColor("#334155")
_SLATE_400  = colors.HexColor("#94A3B8")
_SLATE_200  = colors.HexColor("#E2E8F0")
_SLATE_50   = colors.HexColor("#F8FAFC")
_WHITE      = colors.white


def _fmt_brl(valor: float | None) -> str:
    if valor is None:
        return "—"
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_dt(dt: Any) -> str:
    if dt is None:
        return "—"
    if isinstance(dt, datetime):
        return dt.strftime("%d/%m/%Y %H:%M")
    try:
        d = datetime.fromisoformat(str(dt).replace("Z", "+00:00"))
        return d.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(dt)


def _calcular_duracao(inicio: Any, fim: Any) -> str:
    try:
        if isinstance(inicio, str):
            inicio = datetime.fromisoformat(inicio.replace("Z", "+00:00"))
        if isinstance(fim, str):
            fim = datetime.fromisoformat(fim.replace("Z", "+00:00"))
        if fim is None:
            fim = datetime.now(timezone.utc)
        # Garante que ambos são timezone-aware para subtração segura
        if hasattr(inicio, 'tzinfo') and inicio.tzinfo is None:
            inicio = inicio.replace(tzinfo=timezone.utc)
        if hasattr(fim, 'tzinfo') and fim.tzinfo is None:
            fim = fim.replace(tzinfo=timezone.utc)
        delta = fim - inicio
        h, rem = divmod(int(delta.total_seconds()), 3600)
        m = rem // 60
        if h > 0:
            return f"{h}h {m}min"
        return f"{m}min"
    except Exception:
        return "—"


def gerar_relatorio_final_pdf(
    sessao: Any,
    stats: dict,
    itens: list[dict],
    valor_estoque: dict | None = None,
    analise_ia: dict | None = None,
    historico: list | None = None,
) -> bytes:
    """Gera PDF executivo completo da sessão com análise de erros, acertos e impacto financeiro."""
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"Relatório Final — {sessao.codigo}",
        author="Inventário QR",
    )

    _ss = getSampleStyleSheet()

    def _sty(name: str, **kw) -> ParagraphStyle:
        return ParagraphStyle(name, parent=_ss["Normal"], **kw)

    s_app     = _sty("App", fontSize=9, textColor=_BLUE, fontName="Helvetica-Bold")
    s_title   = _sty("Title", fontSize=20, textColor=_SLATE_900, fontName="Helvetica-Bold", spaceAfter=2*mm)
    s_meta    = _sty("Meta", fontSize=8, textColor=_SLATE_400, spaceAfter=1*mm)
    s_section = _sty("Sec", fontSize=7.5, textColor=_SLATE_400, fontName="Helvetica-Bold",
                     spaceBefore=5*mm, spaceAfter=2*mm, letterSpacing=0.8)
    s_body    = _sty("Body", fontSize=9, textColor=_SLATE_700, leading=14)
    s_footer  = _sty("Footer", fontSize=7, textColor=_SLATE_400, alignment=TA_CENTER)
    s_hdr_tbl = _sty("HdrTbl", fontSize=8, textColor=_WHITE, fontName="Helvetica-Bold", alignment=TA_CENTER)
    s_cell    = _sty("Cell", fontSize=7.5, textColor=_SLATE_700, leading=10)
    s_cell_ok = _sty("CellOk", fontSize=7.5, textColor=_GREEN, fontName="Helvetica-Bold", leading=10, alignment=TA_CENTER)
    s_cell_div= _sty("CellDiv", fontSize=7.5, textColor=_RED, fontName="Helvetica-Bold", leading=10, alignment=TA_CENTER)
    s_cell_pend=_sty("CellPend",fontSize=7.5, textColor=_AMBER, fontName="Helvetica-Bold", leading=10, alignment=TA_CENTER)
    s_cell_c  = _sty("CellC", fontSize=7.5, textColor=_SLATE_700, leading=10, alignment=TA_CENTER)
    s_code    = _sty("Code", fontSize=7, textColor=_BLUE_DARK, fontName="Helvetica-Bold", leading=10)

    story: list[Any] = []

    total = stats.get("total", 0)
    conferidos = stats.get("conferidos", 0)
    pendentes = stats.get("pendentes", 0)
    divergencias = stats.get("divergencias", 0)
    ok_count = conferidos - divergencias
    pct = stats.get("percentual", 0)

    # ── Capa / Header ─────────────────────────────────────────────────────────
    story.append(Paragraph("INVENTÁRIO QR — RELATÓRIO FINAL", s_app))
    story.append(Paragraph(sessao.nome, s_title))
    story.append(Paragraph(
        f"Sessão: <b>{sessao.codigo}</b> &nbsp;·&nbsp; "
        f"Início: <b>{_fmt_dt(sessao.data_inicio)}</b> &nbsp;·&nbsp; "
        f"Encerramento: <b>{_fmt_dt(sessao.data_fim)}</b> &nbsp;·&nbsp; "
        f"Duração: <b>{_calcular_duracao(sessao.data_inicio, sessao.data_fim)}</b>",
        s_meta,
    ))
    story.append(Spacer(1, 3 * mm))
    story.append(HRFlowable(width="100%", thickness=2, color=_BLUE, spaceAfter=5*mm))

    # ── KPIs principais ───────────────────────────────────────────────────────
    story.append(Paragraph("RESUMO EXECUTIVO", s_section))

    acerto_pct = round((ok_count / total) * 100, 1) if total > 0 else 0
    erro_pct   = round((divergencias / total) * 100, 1) if total > 0 else 0

    kpi_data = [
        ["Total de Itens", "Conferidos", "OK (Acertos)", "Divergências", "Progresso"],
        [
            str(total),
            str(conferidos),
            f"{ok_count} ({acerto_pct}%)",
            f"{divergencias} ({erro_pct}%)",
            f"{pct:.1f}%",
        ],
    ]
    kpi_table = Table(kpi_data, colWidths=[36*mm]*5)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), _BLUE),
        ("TEXTCOLOR",  (0,0), (-1,0), _WHITE),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,0), 8),
        ("ALIGN",      (0,0), (-1,0), "CENTER"),
        ("TOPPADDING", (0,0), (-1,0), 4),
        ("BOTTOMPADDING",(0,0),(-1,0),4),
        ("BACKGROUND", (0,1), (-1,1), _SLATE_50),
        ("FONTNAME",   (0,1), (-1,1), "Helvetica-Bold"),
        ("FONTSIZE",   (0,1), (-1,1), 14),
        ("ALIGN",      (0,1), (-1,1), "CENTER"),
        ("TEXTCOLOR",  (2,1), (2,1),  _GREEN),
        ("TEXTCOLOR",  (3,1), (3,1),  _RED),
        ("TEXTCOLOR",  (4,1), (4,1),  _BLUE),
        ("TOPPADDING", (0,1), (-1,1), 6),
        ("BOTTOMPADDING",(0,1),(-1,1),6),
        ("BOX",        (0,0), (-1,-1), 0.5, _SLATE_200),
        ("INNERGRID",  (0,0), (-1,-1), 0.5, _SLATE_200),
    ]))
    story.append(kpi_table)

    # ── Impacto Financeiro ────────────────────────────────────────────────────
    if valor_estoque and valor_estoque.get("tem_dados_financeiros"):
        story.append(Paragraph("IMPACTO FINANCEIRO", s_section))

        vi = valor_estoque.get("valor_inicial", 0) or 0
        vf = valor_estoque.get("valor_final", 0) or 0
        diff_val = valor_estoque.get("diferenca", 0) or 0
        diff_pct = valor_estoque.get("percentual_variacao", 0) or 0
        sinal = "+" if diff_val >= 0 else ""
        diff_color = _GREEN if diff_val >= 0 else _RED

        fin_data = [
            ["Valor Inicial (Base)", "Valor Apurado (Contagem)", "Variação", "Variação %"],
            [_fmt_brl(vi), _fmt_brl(vf), f"{sinal}{_fmt_brl(diff_val)}", f"{sinal}{diff_pct:.2f}%"],
        ]
        fin_table = Table(fin_data, colWidths=[45*mm]*4)
        fin_table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), _SLATE_900),
            ("TEXTCOLOR",  (0,0), (-1,0), _WHITE),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,0), 8),
            ("ALIGN",      (0,0), (-1,0), "CENTER"),
            ("TOPPADDING", (0,0), (-1,0), 4),
            ("BOTTOMPADDING",(0,0),(-1,0),4),
            ("BACKGROUND", (0,1), (-1,1), _SLATE_50),
            ("FONTNAME",   (0,1), (-1,1), "Helvetica-Bold"),
            ("FONTSIZE",   (0,1), (-1,1), 13),
            ("ALIGN",      (0,1), (-1,1), "CENTER"),
            ("TEXTCOLOR",  (2,1), (2,1),  diff_color),
            ("TEXTCOLOR",  (3,1), (3,1),  diff_color),
            ("TOPPADDING", (0,1), (-1,1), 6),
            ("BOTTOMPADDING",(0,1),(-1,1),6),
            ("BOX",        (0,0), (-1,-1), 0.5, _SLATE_200),
            ("INNERGRID",  (0,0), (-1,-1), 0.5, _SLATE_200),
        ]))
        story.append(fin_table)
        story.append(Spacer(1, 2*mm))

        # Top perdas e ganhos
        perdas = valor_estoque.get("maiores_perdas", [])[:5]
        ganhos = valor_estoque.get("maiores_ganhos", [])[:5]

        if perdas or ganhos:
            def _tabela_top(titulo, items_list, cor):
                rows_top = [[Paragraph(titulo, _sty(f"T{titulo}", fontSize=8, textColor=cor, fontName="Helvetica-Bold"))]]
                for it in items_list:
                    rows_top.append([Paragraph(
                        f"<b>{it.get('codigo','—')}</b> — {it.get('produto','')} &nbsp;"
                        f"<font color='{'#16A34A' if cor==_GREEN else '#DC2626'}'>"
                        f"{'+'if cor==_GREEN else ''}{_fmt_brl(it.get('diferenca_valor'))}</font>",
                        _sty(f"R{titulo}", fontSize=7.5, textColor=_SLATE_700, leading=11)
                    )])
                t = Table(rows_top, colWidths=[85*mm])
                t.setStyle(TableStyle([
                    ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#F8FAFC")),
                    ("TOPPADDING", (0,0), (-1,-1), 3),
                    ("BOTTOMPADDING",(0,0),(-1,-1),3),
                    ("LEFTPADDING",(0,0),(-1,-1),6),
                    ("RIGHTPADDING",(0,0),(-1,-1),6),
                    ("BOX", (0,0),(-1,-1),0.5,_SLATE_200),
                    ("INNERGRID",(0,0),(-1,-1),0.25,_SLATE_200),
                ]))
                return t

            top_row = [[_tabela_top("Maiores Perdas", perdas, _RED),
                        _tabela_top("Maiores Ganhos", ganhos, _GREEN)]]
            top_table = Table(top_row, colWidths=[90*mm, 90*mm])
            top_table.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
            story.append(top_table)

    # ── Análise IA ────────────────────────────────────────────────────────────
    if analise_ia and (analise_ia.get("resumo") or analise_ia.get("recomendacoes")):
        story.append(Paragraph("ANÁLISE DE INTELIGÊNCIA ARTIFICIAL", s_section))

        resumo = analise_ia.get("resumo") or analise_ia.get("relatorio_executivo") or ""
        if resumo:
            resumo_block = [
                [Paragraph("Resumo Executivo", _sty("ResHdr", fontSize=8, textColor=_PURPLE, fontName="Helvetica-Bold"))],
                [Paragraph(resumo, s_body)],
            ]
            rt = Table(resumo_block, colWidths=[180*mm])
            rt.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0), _PURPLE_LIGHT),
                ("BACKGROUND",(0,1),(-1,-1), colors.HexColor("#FAF9FF")),
                ("TOPPADDING",(0,0),(-1,-1),5),
                ("BOTTOMPADDING",(0,0),(-1,-1),5),
                ("LEFTPADDING",(0,0),(-1,-1),8),
                ("RIGHTPADDING",(0,0),(-1,-1),8),
                ("BOX",(0,0),(-1,-1),0.5, colors.HexColor("#DDD6FE")),
                ("INNERGRID",(0,0),(-1,-1),0.25,_SLATE_200),
            ]))
            story.append(rt)
            story.append(Spacer(1, 2*mm))

        recs = analise_ia.get("recomendacoes") or []
        if recs:
            story.append(Paragraph("Recomendações de Melhoria", _sty("RecHdr", fontSize=9, textColor=_SLATE_700, fontName="Helvetica-Bold", spaceBefore=2*mm, spaceAfter=1*mm)))
            for i, rec in enumerate(recs[:10], 1):
                text = rec if isinstance(rec, str) else (rec.get("texto") or rec.get("text") or str(rec))
                story.append(Paragraph(
                    f"<b>{i}.</b> {text}",
                    _sty(f"Rec{i}", fontSize=8.5, textColor=_SLATE_700, leading=13, leftIndent=8)
                ))
                story.append(Spacer(1, 1*mm))

    # ── Tabela completa de itens ───────────────────────────────────────────────
    story.append(Paragraph(f"DETALHAMENTO DE ITENS — {len(itens)} registros", s_section))

    col_widths = [22*mm, 52*mm, 13*mm, 18*mm, 14*mm, 20*mm, 24*mm, 17*mm]
    hdr = [
        Paragraph("Código", s_hdr_tbl),
        Paragraph("Produto", s_hdr_tbl),
        Paragraph("Base", s_hdr_tbl),
        Paragraph("Contado", s_hdr_tbl),
        Paragraph("Dif.", s_hdr_tbl),
        Paragraph("Status", s_hdr_tbl),
        Paragraph("Operador", s_hdr_tbl),
        Paragraph("Rodada", s_hdr_tbl),
    ]
    rows: list[list] = [hdr]
    for item in itens:
        diff = item.get("diferenca")
        status = str(item.get("status", ""))
        diff_str = "—" if diff is None else (f"+{diff}" if diff > 0 else str(diff))
        if status == "Divergente":
            sp = Paragraph(status, s_cell_div)
        elif status == "Pendente":
            sp = Paragraph(status, s_cell_pend)
        else:
            sp = Paragraph(status, s_cell_ok)
        rows.append([
            Paragraph(str(item.get("codigo", "")), s_code),
            Paragraph(str(item.get("produto", "")), s_cell),
            Paragraph(str(item.get("quantidade_base", "")), s_cell_c),
            Paragraph("—" if item.get("quantidade_encontrada") is None else str(item["quantidade_encontrada"]), s_cell_c),
            Paragraph(diff_str, s_cell_c),
            sp,
            Paragraph(str(item.get("operador") or "—"), s_cell),
            Paragraph(str(item.get("rodada") or "—"), s_cell_c),
        ])

    items_table = Table(rows, colWidths=col_widths, repeatRows=1)
    cmds_it = [
        ("BACKGROUND",    (0,0), (-1,0),  _BLUE),
        ("TOPPADDING",    (0,0), (-1,0),  4),
        ("BOTTOMPADDING", (0,0), (-1,0),  4),
        ("LEFTPADDING",   (0,0), (-1,-1), 4),
        ("RIGHTPADDING",  (0,0), (-1,-1), 4),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,1), (-1,-1), 3),
        ("BOTTOMPADDING", (0,1), (-1,-1), 3),
        ("BOX",           (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ("INNERGRID",     (0,0), (-1,-1), 0.25, _SLATE_200),
    ]
    for i, item in enumerate(itens):
        r = i + 1
        st = item.get("status", "")
        if st == "Divergente":
            cmds_it.append(("BACKGROUND", (0,r), (-1,r), _RED_LIGHT))
        elif st == "Pendente":
            cmds_it.append(("BACKGROUND", (0,r), (-1,r), _AMBER_LIGHT))
        elif i % 2 == 1:
            cmds_it.append(("BACKGROUND", (0,r), (-1,r), _SLATE_50))
    items_table.setStyle(TableStyle(cmds_it))
    story.append(items_table)

    # ── Histórico de Rodadas ─────────────────────────────────────────────────
    if historico:
        story.append(Spacer(1, 6*mm))
        story.append(Paragraph("HISTÓRICO DE RODADAS POR ITEM", s_section))
        story.append(Spacer(1, 2*mm))
        story.append(Paragraph(
            "Itens que precisaram de mais de 1 rodada ou finalizaram como Para Ajuste.",
            s_body,
        ))
        story.append(Spacer(1, 3*mm))

        # Agrupa por código
        from collections import defaultdict
        hist_por_item: dict[str, list] = defaultdict(list)
        for h in historico:
            cod = getattr(h, "codigo", None) or (h.get("codigo") if isinstance(h, dict) else None)
            if cod:
                hist_por_item[cod].append(h)

        # Filtra apenas itens com mais de 1 contagem ou para_ajuste
        itens_multi = {
            cod: regs for cod, regs in hist_por_item.items()
            if len(regs) > 1 or any(
                (getattr(r, "para_ajuste", False) or (isinstance(r, dict) and r.get("para_ajuste")))
                for r in regs
            )
        }

        if itens_multi:
            rod_cols = ["Código", "Produto", "Qtd. Contagens", "Rodadas", "Resultado", "Operador(es)"]
            rod_widths = [30*mm, 60*mm, 25*mm, 22*mm, 28*mm, 35*mm]
            rod_rows = [rod_cols]
            for cod, regs in sorted(itens_multi.items()):
                regs_sorted = sorted(regs, key=lambda r: getattr(r, "timestamp", None) or (r.get("timestamp") if isinstance(r, dict) else ""))
                rodadas = sorted({getattr(r, "rodada", 1) or (r.get("rodada") if isinstance(r, dict) else 1) for r in regs_sorted})
                operadores = list({str(getattr(r, "operador", "") or (r.get("operador") if isinstance(r, dict) else "") or "—") for r in regs_sorted})
                ultimo = regs_sorted[-1]
                pa = getattr(ultimo, "para_ajuste", False) or (isinstance(ultimo, dict) and ultimo.get("para_ajuste"))
                div = getattr(ultimo, "divergencia", False) or (isinstance(ultimo, dict) and ultimo.get("divergencia"))
                produto = next((it.get("produto") for it in itens if it.get("codigo") == cod), "—")
                resultado = "Para Ajuste" if pa else ("Divergente" if div else "OK")
                rod_rows.append([
                    Paragraph(cod, s_cell),
                    Paragraph(str(produto), s_cell),
                    Paragraph(str(len(regs)), s_cell_c),
                    Paragraph(", ".join(str(r) for r in rodadas), s_cell_c),
                    Paragraph(resultado, s_cell_c),
                    Paragraph(", ".join(operadores[:3]), s_cell),
                ])

            rod_table = Table(rod_rows, colWidths=rod_widths, repeatRows=1)
            rod_cmds = [
                ("BACKGROUND",  (0, 0), (-1, 0), _BLUE),
                ("TEXTCOLOR",   (0, 0), (-1, 0), _WHITE),
                ("FONTSIZE",    (0, 0), (-1, 0), 8),
                ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",    (0, 1), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_WHITE, _SLATE_50]),
                ("BOX",         (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("INNERGRID",   (0, 0), (-1, -1), 0.25, _SLATE_200),
                ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING",  (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
            for i, (cod, regs) in enumerate(sorted(itens_multi.items()), 1):
                ultimo = sorted(regs, key=lambda r: getattr(r, "timestamp", None) or "")[-1]
                pa = getattr(ultimo, "para_ajuste", False) or (isinstance(ultimo, dict) and ultimo.get("para_ajuste"))
                div = getattr(ultimo, "divergencia", False) or (isinstance(ultimo, dict) and ultimo.get("divergencia"))
                if pa:
                    rod_cmds.append(("BACKGROUND", (0, i), (-1, i), _PURPLE_LIGHT))
                elif div:
                    rod_cmds.append(("BACKGROUND", (0, i), (-1, i), _RED_LIGHT))
            rod_table.setStyle(TableStyle(rod_cmds))
            story.append(rod_table)
        else:
            story.append(Paragraph("Todos os itens foram conferidos na 1ª rodada sem divergências.", s_body))

    # ── Footer ────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 5*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=_SLATE_200, spaceAfter=2*mm))
    now = datetime.now(timezone.utc).strftime("%d/%m/%Y às %H:%M UTC")
    story.append(Paragraph(f"Relatório Final gerado em {now} · Inventário QR · {sessao.codigo}", s_footer))

    doc.build(story)
    return buf.getvalue()


def gerar_relatorio_final_excel(
    sessao: Any,
    stats: dict,
    itens: list[dict],
    valor_estoque: dict | None = None,
    analise_ia: dict | None = None,
    historico: list | None = None,
    metricas: dict | None = None,
) -> bytes:
    """Gera Excel com múltiplas abas: Resumo, Itens OK, Divergências, Impacto Financeiro, Recomendações."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output = BytesIO()

    total = stats.get("total", 0)
    conferidos = stats.get("conferidos", 0)
    pendentes = stats.get("pendentes", 0)
    divergencias = stats.get("divergencias", 0)
    ok_count = conferidos - divergencias
    pct = stats.get("percentual", 0)
    acerto_pct = round((ok_count / total) * 100, 1) if total > 0 else 0

    # ── Aba 1: Resumo Executivo ────────────────────────────────────────────────
    resumo_data = {
        "Sessão": [sessao.nome],
        "Código": [sessao.codigo],
        "Início": [_fmt_dt(sessao.data_inicio)],
        "Encerramento": [_fmt_dt(sessao.data_fim)],
        "Duração": [_calcular_duracao(sessao.data_inicio, sessao.data_fim)],
        "Total de Itens": [total],
        "Itens Conferidos": [conferidos],
        "Itens OK (Acertos)": [ok_count],
        "Taxa de Acerto (%)": [acerto_pct],
        "Divergências": [divergencias],
        "Taxa de Divergência (%)": [round((divergencias/total)*100,1) if total > 0 else 0],
        "Itens Pendentes": [pendentes],
        "Progresso (%)": [round(pct, 1)],
    }

    if valor_estoque and valor_estoque.get("tem_dados_financeiros"):
        resumo_data["Valor Inicial (R$)"] = [valor_estoque.get("valor_inicial")]
        resumo_data["Valor Apurado (R$)"] = [valor_estoque.get("valor_final")]
        resumo_data["Variação (R$)"] = [valor_estoque.get("diferenca")]
        resumo_data["Variação (%)"] = [round(valor_estoque.get("percentual_variacao", 0), 2)]

    df_resumo = pd.DataFrame(resumo_data).T.reset_index()
    df_resumo.columns = ["Campo", "Valor"]

    # ── Aba 2: Todos os Itens ─────────────────────────────────────────────────
    cols_itens = ["codigo", "produto", "quantidade_base", "quantidade_encontrada", "diferenca",
                  "status", "operador", "rodada", "observacao", "local_fisico", "timestamp"]
    df_itens = pd.DataFrame(itens).reindex(columns=cols_itens)
    df_itens.columns = ["Código", "Produto", "Base", "Contado", "Diferença", "Status",
                        "Operador", "Rodada", "Observação", "Local", "Data/Hora"]

    # ── Aba 3: Divergências + Para Ajuste ────────────────────────────────────
    divs = [it for it in itens if str(it.get("status", "")).lower() in ("divergente", "para ajuste")]
    cols_divs = cols_itens + ["diferenca_valor"]
    df_divs = pd.DataFrame(divs).reindex(columns=cols_divs) if divs else pd.DataFrame(columns=cols_divs)
    if not df_divs.empty:
        # "diferenca_valor" já vem populado no dict de cada item (de montar_inventario_completo).
        # NÃO usar maiores_perdas/maiores_ganhos que são truncados ao top-5 — isso deixaria
        # NaN para todos os outros itens divergentes.
        # O campo já está em cols_divs; apenas garante que a série existe.
        if "diferenca_valor" not in df_divs.columns:
            df_divs["diferenca_valor"] = None
        df_divs.columns = ["Código", "Produto", "Base", "Contado", "Diferença", "Status",
                           "Operador", "Rodada", "Observação", "Local", "Data/Hora", "Impacto (R$)"]

    # ── Aba 4: Recomendações ──────────────────────────────────────────────────
    recs = []
    if analise_ia:
        for i, r in enumerate((analise_ia.get("recomendacoes") or [])[:20], 1):
            text = r if isinstance(r, str) else (r.get("texto") or r.get("text") or str(r))
            recs.append({"#": i, "Recomendação": text, "Origem": "IA"})
    # Recomendações padrão baseadas nos dados
    if divergencias > 0:
        pct_div = round((divergencias / total) * 100, 1) if total > 0 else 0
        recs.append({"#": len(recs)+1, "Recomendação": f"Revisar {divergencias} itens divergentes ({pct_div}% do total) antes de fechar o estoque.", "Origem": "Sistema"})
    if pendentes > 0:
        recs.append({"#": len(recs)+1, "Recomendação": f"Há {pendentes} item(ns) sem contagem registrada. Verifique se foram omitidos.", "Origem": "Sistema"})
    if not recs:
        recs.append({"#": 1, "Recomendação": "Inventário concluído sem pendências significativas.", "Origem": "Sistema"})
    df_recs = pd.DataFrame(recs)

    # ── Aba 5: Histórico de Rodadas ───────────────────────────────────────────
    df_hist = None
    if historico:
        hist_rows = []
        for h in historico:
            if isinstance(h, dict):
                hist_rows.append(h)
            else:
                hist_rows.append({
                    "codigo": getattr(h, "codigo", ""),
                    "rodada": getattr(h, "rodada", 1),
                    "quantidade_encontrada": getattr(h, "quantidade_encontrada", None),
                    "quantidade_base": getattr(h, "quantidade_base", None),
                    "divergencia": getattr(h, "divergencia", False),
                    "para_ajuste": getattr(h, "para_ajuste", False),
                    "operador": getattr(h, "operador", ""),
                    "observacao": getattr(h, "observacao", ""),
                    "timestamp": getattr(h, "timestamp", None),
                })
        if hist_rows:
            df_hist = pd.DataFrame(hist_rows).reindex(
                columns=["codigo", "rodada", "quantidade_encontrada", "quantidade_base",
                         "divergencia", "para_ajuste", "operador", "observacao", "timestamp"]
            )
            df_hist.columns = ["Código", "Rodada", "Qtd. Encontrada", "Qtd. Base",
                                "Divergente", "Para Ajuste", "Operador", "Observação", "Data/Hora"]
            # Linha de resumo por item: quantas contagens, resultado final
            prod_map = {it.get("codigo"): it.get("produto", "") for it in itens}
            resumo_hist = (
                df_hist.groupby("Código")
                .agg(
                    Produto=("Código", lambda x: prod_map.get(x.iloc[0], "")),
                    Contagens=("Rodada", "count"),
                    Rodadas=("Rodada", lambda x: ", ".join(str(r) for r in sorted(x.unique()))),
                    Resultado_Final=("Para Ajuste", lambda x: "Para Ajuste" if x.iloc[-1] else ("Divergente" if df_hist.loc[x.index[-1:], "Divergente"].iloc[0] else "OK")),
                    Operadores=("Operador", lambda x: ", ".join(str(v) for v in x.dropna().unique() if v)[:50]),
                )
                .reset_index()
            )
            resumo_hist.columns = ["Código", "Produto", "Total Contagens", "Rodadas", "Resultado Final", "Operadores"]

    # ── Aba de Métricas de Produtividade ──────────────────────────────────────
    df_metricas_resumo = None
    df_metricas_por_op = None
    if metricas:
        met_rows = [
            ("Duração Total (min)", metricas.get("duracao_minutos")),
            ("Total de Itens", metricas.get("total_itens")),
            ("Total de Contagens (estado atual)", metricas.get("total_contagens_atuais")),
            ("Total de Tentativas (histórico)", metricas.get("total_tentativas_historico")),
            ("Itens por Minuto (geral)", metricas.get("itens_por_minuto")),
            ("Divergências (absoluto)", metricas.get("divergencias_absolutas")),
            ("Taxa de Divergência (%)", metricas.get("taxa_divergencia_pct")),
            ("Retrabalho (tentativas extras)", metricas.get("retrabalho_absoluto")),
            ("Taxa de Retrabalho (%)", metricas.get("taxa_retrabalho_pct")),
            ("Contagens com Operador", metricas.get("contagens_com_operador")),
            ("% Rastreabilidade (operador+timestamp)", metricas.get("pct_rastreabilidade")),
        ]
        df_metricas_resumo = pd.DataFrame(met_rows, columns=["Métrica", "Valor"])

        por_op = metricas.get("por_operador") or []
        if por_op:
            df_metricas_por_op = pd.DataFrame(por_op).reindex(
                columns=["operador", "contagens", "itens_unicos",
                         "primeiro_registro", "ultimo_registro",
                         "duracao_minutos", "itens_por_minuto"]
            )
            df_metricas_por_op.columns = [
                "Operador", "Tentativas", "Itens Únicos",
                "Primeiro Registro", "Último Registro",
                "Duração (min)", "Itens/min",
            ]

    # ── Escreve para Excel com formatação ─────────────────────────────────────
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_resumo.to_excel(writer, index=False, sheet_name="Resumo Executivo")
        df_itens.to_excel(writer, index=False, sheet_name="Todos os Itens")
        if not df_divs.empty:
            df_divs.to_excel(writer, index=False, sheet_name="Divergências")
        df_recs.to_excel(writer, index=False, sheet_name="Recomendações")
        if df_metricas_resumo is not None:
            df_metricas_resumo.to_excel(writer, index=False, sheet_name="Métricas Produtividade")
        if df_metricas_por_op is not None:
            df_metricas_por_op.to_excel(writer, index=False, sheet_name="Produtividade por Operador")
        if df_hist is not None:
            resumo_hist.to_excel(writer, index=False, sheet_name="Resumo por Rodadas")
            df_hist.to_excel(writer, index=False, sheet_name="Histórico Detalhado")

        wb = writer.book

        # Estilos reutilizáveis
        hdr_fill   = PatternFill("solid", fgColor="2563EB")
        hdr_font   = Font(bold=True, color="FFFFFF", size=10)
        ok_fill    = PatternFill("solid", fgColor="F0FDF4")
        div_fill   = PatternFill("solid", fgColor="FEF2F2")
        pend_fill  = PatternFill("solid", fgColor="FFFBEB")
        alt_fill   = PatternFill("solid", fgColor="F8FAFC")
        thin_border= Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        def _estilizar_aba(ws, larguras: list[int]):
            for i, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
                for cell in col:
                    cell.fill = hdr_fill
                    cell.font = hdr_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin_border
            ws.row_dimensions[1].height = 28
            for r_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
            for i, w in enumerate(larguras, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        # Aba Resumo
        ws_res = wb["Resumo Executivo"]
        _estilizar_aba(ws_res, [30, 25])
        ws_res.column_dimensions["A"].width = 30
        ws_res.column_dimensions["B"].width = 25

        # Aba Todos os Itens
        ws_it = wb["Todos os Itens"]
        _estilizar_aba(ws_it, [18, 40, 10, 10, 10, 14, 18, 10, 30, 20, 20])
        status_col = 6
        for row in ws_it.iter_rows(min_row=2):
            st = str(row[status_col - 1].value or "").lower()
            fill = ok_fill if st == "ok" else div_fill if st == "divergente" else pend_fill if st == "pendente" else (alt_fill if row[0].row % 2 == 0 else None)
            if fill:
                for cell in row:
                    cell.fill = fill

        # Aba Divergências
        if "Divergências" in wb.sheetnames:
            ws_div = wb["Divergências"]
            _estilizar_aba(ws_div, [18, 40, 10, 10, 10, 16, 18, 10, 30, 20, 20, 16])
            for row in ws_div.iter_rows(min_row=2):
                st_cell = row[5].value or ""
                fill = div_fill if str(st_cell).lower() == "divergente" else PatternFill("solid", fgColor="F5F3FF")
                for cell in row:
                    cell.fill = fill

        # Aba Recomendações
        ws_rec = wb["Recomendações"]
        _estilizar_aba(ws_rec, [6, 80, 14])

        # Abas de Métricas de Produtividade
        if df_metricas_resumo is not None and "Métricas Produtividade" in wb.sheetnames:
            ws_met = wb["Métricas Produtividade"]
            _estilizar_aba(ws_met, [40, 20])
        if df_metricas_por_op is not None and "Produtividade por Operador" in wb.sheetnames:
            ws_op = wb["Produtividade por Operador"]
            _estilizar_aba(ws_op, [25, 14, 14, 22, 22, 16, 12])

        # Abas de Histórico de Rodadas
        if df_hist is not None:
            ws_res_hist = wb["Resumo por Rodadas"]
            _estilizar_aba(ws_res_hist, [18, 40, 18, 20, 18, 35])
            # Colorir por resultado
            res_col = 5
            pur_fill = PatternFill("solid", fgColor="F5F3FF")
            for row in ws_res_hist.iter_rows(min_row=2):
                res_val = str(row[res_col - 1].value or "").lower()
                if "ajuste" in res_val:
                    f = pur_fill
                elif "divergente" in res_val:
                    f = div_fill
                else:
                    f = ok_fill
                for cell in row:
                    cell.fill = f
            ws_hist = wb["Histórico Detalhado"]
            _estilizar_aba(ws_hist, [18, 10, 16, 12, 12, 12, 20, 30, 20])

    return output.getvalue()
